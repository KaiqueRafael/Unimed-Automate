import os
import time
import datetime

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

from selenium import webdriver

from selenium.webdriver.support.ui import WebDriverWait

from selenium.webdriver.common.by import By

from selenium.webdriver.common.keys import Keys

from selenium.webdriver.support import expected_conditions as EC


class processoAutorizacao:
    def __init__(self):
        self.navegador = webdriver.Chrome()
        self.aguardar = WebDriverWait(self.navegador, 10)
        self.aguardarErro = WebDriverWait(self.navegador, 2)

    def iniciar(self):
        self.navegador.get("https://wws3.unimedcampinas.com.br/NewSadTISS/")
        self.navegador.maximize_window()

        time.sleep(2)

        self.usu_senha()
        self.navegacao()
        self.valid_guia_intercambio()

    def usu_senha(self):
        self.usuario = '//*[@id="ctl00_ContentPlaceHolder1_cLogin_txtLogin"]'
        self.senha = '//*[@id="ctl00_ContentPlaceHolder1_cLogin_txtSenha"]'
        self.navegador.find_element(By.XPATH, (self.usuario)).send_keys('usuario')
        self.navegador.find_element(By.XPATH, (self.senha)).send_keys('senha' + Keys.ENTER)

    def navegacao(self):
        self.btn_salvar = self.aguardar.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="btnSalvar"]')))
        self.btn_salvar.click()

        self.btn_menu = self.aguardar.until(
            EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]')))
        self.btn_menu.click()
        time.sleep(2)
        self.btn_menu_pesquisa = self.navegador.find_element(
            By.XPATH, '/html/body/div[3]/ul/li[9]')
        self.btn_menu_pesquisa.click()
        self.btn_consulta_situacao = self.navegador.find_element(
            By.XPATH, '/html/body/div[4]/ul/li[2]')
        self.btn_consulta_situacao.click()
        time.sleep(2)

    def criando_cabecalho_xl(self):
        self.novo_excel = Workbook()
        self.inserir = self.novo_excel.active

        self.inserir["A1"] = "Guia"
        self.inserir["B1"] = "Status da Guia"
        self.inserir["C1"] = "Atendimento"
        self.inserir["D1"] = "Paciente"
        self.inserir["E1"] = "Validade da Guia"
        self.inserir["F1"] = "Cod Interno"
        self.inserir["G1"] = "Cod Unimed"
        self.inserir["H1"] = "Status do Cod Procedimento"

        self.fundo = PatternFill(start_color='5399FF',
                                 end_color='5399FF',
                                 fill_type='gray125')

        self.fonte = Font(color="ffffff")

        self.inserir["A1"].fill = self.fundo
        self.inserir['A1'].font = self.fonte
        self.inserir.column_dimensions['A'].width = 20
        self.inserir["B1"].fill = self.fundo
        self.inserir['B1'].font = self.fonte
        self.inserir.column_dimensions['B'].width = 60
        self.inserir["C1"].fill = self.fundo
        self.inserir['C1'].font = self.fonte
        self.inserir.column_dimensions['C'].width = 30
        self.inserir["D1"].fill = self.fundo
        self.inserir['D1'].font = self.fonte
        self.inserir.column_dimensions['D'].width = 60
        self.inserir["E1"].fill = self.fundo
        self.inserir['E1'].font = self.fonte
        self.inserir.column_dimensions['E'].width = 30
        self.inserir["F1"].fill = self.fundo
        self.inserir['F1'].font = self.fonte
        self.inserir.column_dimensions['F'].width = 30
        self.inserir["G1"].fill = self.fundo
        self.inserir['G1'].font = self.fonte
        self.inserir.column_dimensions['G'].width = 50
        self.inserir["H1"].fill = self.fundo
        self.inserir['H1'].font = self.fonte
        self.inserir.column_dimensions['H'].width = 50

        self.data_final = datetime.datetime.now()

        self.novo_excel.save(f"local/base.xlsx")

    def valid_guia_intercambio(self):

        self.lerArquivo = "arquivo.xlsx"
        self.arqExcel = pd.read_excel(self.lerArquivo)

        # ==== Criando nova planilha através da função ====
        
        self.criando_cabecalho_xl()

        for i, n_guia in enumerate(self.arqExcel["Guia"]):

            procedimento = self.arqExcel["Procedimento"][i]
            atendimento = self.arqExcel["Atendimento"][i]

            self.insere_guia = self.navegador.find_element(By.XPATH,
                                                           '/html/body/div[2]/form/div[1]/div/div[1]/div[2]/input')
            self.insere_guia.click()

            def guiaIntercambio(intercambio):
                rec_guia = intercambio
                edita_qtd_guia = str(rec_guia)[:9]
                return edita_qtd_guia

            def tratandoGuia(info_guia):
                qtd_caract_guia = str(info_guia).startswith('13')

                if qtd_caract_guia is True:
                    return guiaIntercambio(info_guia)
                else:
                    return info_guia

            self.guia = tratandoGuia(str(n_guia).replace(" ", ""))

            # ==== Consultando guia pelo valor tratado ====

            self.insere_guia.send_keys(self.guia, Keys.ENTER)

            try:
                alerta = self.aguardarErro.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="popup_ok"]')))
                time.sleep(1)
                alerta.click()
                status = 'ERRO'
            except:
                status = self.navegador.find_element(By.XPATH, '//*[@id="dadosSolicitacao_status"]').text
                paciente = self.navegador.find_element(By.XPATH, '//*[@id="dadosSolicitacao_nomeBeneficiario"]').text
                validade_guia = self.navegador.find_element(By.XPATH, '//*[@id="dadosSolicitacao_dtValidade"]').text
                proc_unimed = self.navegador.find_element(By.XPATH, '//*[@id="GridItensSolicProcedimento"]/tbody/tr/td[1]').text

            if status == 'ERRO':
                status = 'NÚMERO DE GUIA INVÁLIDO'
                paciente = ''
            else:
                status = status

            if str(procedimento)[:8] == proc_unimed:
                sts_procedimento = 'Mesmo código TUSS'
            elif str(procedimento)[:8] == 'nan':
                sts_procedimento = ''
                procedimento = ''
            else:
                sts_procedimento = 'Código TUSS diferente'

            # ==== Inserindo informações para o novo arquivo excel ====
            linha = 2
            linha += i
            self.inserir[f"A{linha}"] = str(self.guia)
            self.inserir[f"B{linha}"] = status
            self.inserir[f"C{linha}"] = str(atendimento)[:-2]
            self.inserir[f"D{linha}"] = paciente
            self.inserir[f"E{linha}"] = str(validade_guia)[:10]
            self.inserir[f"F{linha}"] = str(procedimento)
            self.inserir[f"G{linha}"] = proc_unimed
            self.inserir[f"H{linha}"] = sts_procedimento


            print(linha, ' - ', self.guia, ' - ', status, ' - ', paciente, ' ->', sts_procedimento)

            self.limpa_campo = self.navegador.find_element(By.XPATH,
                                                           '/html/body/div[2]/form/div[1]/div/div[1]/div[2]/input')
            self.limpa_campo.clear()

            time.sleep(1)

        self.novo_excel.save(f"local/base.xlsx")

bot = processoAutorizacao()
bot.iniciar()
